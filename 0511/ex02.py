# -*- coding: utf-8 -*-
"""
Created on Thu May 11 19:04:31 2023

@author: USER
"""

import openpyxl #excel函式庫

file = 'school.xlsx'
wb = openpyxl.load_workbook(file)
#抓取預設工作表
ws = wb.active
print('目前工作表:',ws.title)
ws = wb['students']
print('所選工作表名稱:',ws.title)

print('A2儲存格值:',ws['A2'].value)
print('B2儲存格值:',ws['B2'].value)
print('C2儲存格值:',ws['C2'].value)
print()
print('A2是第',ws['A2'].column,'欄','第',ws['A2'].row,'列') #column是第幾欄，row是第幾列
print('C1是第',ws['C1'].column,'欄','第',ws['C1'].row,'列')

print('工作表欄數:',ws.max_column)
print('工作表列數:',ws.max_row)