# -*- coding: utf-8 -*-
"""
Created on Thu May 11 19:48:40 2023

@author: USER
"""

import openpyxl #excel函式庫

file = 'school.xlsx'
wb = openpyxl.load_workbook(file)
#抓取預設工作表
#ws = wb.active
#print('目前工作表:',ws.title)
ws = wb['students']
print('所選工作表名稱:',ws.title)
for i in range(1,ws.max_column+1):
    print(ws.cell(column=i,row=1).value,end=' ')
print()
for i in range(1,ws.max_column+1):
    print(ws.cell(column=i,row=5).value,end=' ')
print('\n')
for i in range(1,ws.max_row+1):
    for j in range(1,ws.max_column+1):
        print(ws.cell(i,j).value,end=' ')
    print()

#第二種寫法
for row in ws.rows:
    for cell in row: #cell是儲存格
        print(cell.value,end=' ')
    print()