# -*- coding: utf-8 -*-
"""
Created on Thu May 11 20:47:07 2023

@author: USER
"""

import openpyxl

wb = openpyxl.Workbook() #建立空白活頁簿
ws = wb.active

ws.title = 'member' #定義活頁簿名稱

ws['A1']='姓名'
ws['A2']='Bill'
ws['B1']='學校'
ws['B2']='聯成'

rowdata = ['David','中科大']
ws.append(rowdata) #新增一列

#利用迴圈搭配二維陣列新增三列
data = [['John','中興'],['Mary','靜宜'],['Sue','東海']]
for item in data:
    ws.append(item)

wb.save('lcc2.xlsx')